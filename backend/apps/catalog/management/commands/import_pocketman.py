"""포켓만프로젝트 엑셀 → ops-catalog DB import (1회).

원본: archive/포켓만프로젝트/[리더십] 포켓만프로젝트 우선순위.xlsx (시트: 교육운영실)
HM-26 D3: 1회 import 후 시스템이 SoT.

사용:
    python manage.py import_pocketman --xlsx /path/to/file.xlsx [--dry-run]
"""

from __future__ import annotations

from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.catalog.models import (
    Category,
    Difficulty,
    Priority,
    Project,
    Stage,
    Tier,
)

# HM-25 에서 확정한 카테고리 정의 (8개) + 행 → 카테고리 매핑
CATEGORIES = [
    {"code": "T1·1", "tier": Tier.T1, "num": 1, "title": "공지·모니터링 자동화",
     "note": "정기 시점 데이터·공지 자동 발송"},
    {"code": "T1·2", "tier": Tier.T1, "num": 2, "title": "수강생 트래킹·면담·출결",
     "note": "수강생 상태 자동 추적·기록"},
    {"code": "T1·3", "tier": Tier.T1, "num": 3, "title": "평가·지표 분석",
     "note": "다면평가·만족도·KDT 지표 분석"},
    {"code": "T1·5", "tier": Tier.T1, "num": 5, "title": "수강생 개인화·앱",
     "note": "개인화 분석·취업 지원·수강생 앱"},
    {"code": "T2·4", "tier": Tier.T2, "num": 4, "title": "커리큘럼·교안 콘텐츠",
     "note": "커리큘럼 검수·제안 + 교안/PPT 변환"},
    {"code": "T2·7", "tier": Tier.T2, "num": 7, "title": "지식 적재·어시스턴트",
     "note": "맥락 적재 + 검색·요약 + 사내 양식"},
    {"code": "T3·6", "tier": Tier.T3, "num": 6, "title": "업무·CTA 트래킹",
     "note": "프로젝트 진행·리마인드·대시보드"},
    {"code": "T3·8", "tier": Tier.T3, "num": 8, "title": "행정·CS 보조",
     "note": "행정서류·비용 검수·CS·윤문"},
]

# 엑셀 행 번호 (R2~R47) → 카테고리 code
ROW_TO_CAT = {
    # T1·1 공지·모니터링
    4: "T1·1", 17: "T1·1", 21: "T1·1", 47: "T1·1", 19: "T1·1", 28: "T1·1",
    # T1·2 수강생 트래킹
    5: "T1·2", 6: "T1·2", 13: "T1·2", 31: "T1·2", 35: "T1·2", 43: "T1·2", 15: "T1·2", 18: "T1·2",
    # T1·3 평가·지표
    14: "T1·3", 26: "T1·3", 27: "T1·3", 30: "T1·3",
    # T1·5 수강생 개인화·앱
    11: "T1·5", 12: "T1·5", 23: "T1·5", 41: "T1·5",
    # T2·4 커리큘럼·교안
    8: "T2·4", 10: "T2·4", 34: "T2·4", 38: "T2·4", 45: "T2·4", 46: "T2·4",
    # T2·7 지식 적재
    2: "T2·7", 3: "T2·7", 7: "T2·7", 9: "T2·7", 20: "T2·7", 24: "T2·7", 39: "T2·7", 40: "T2·7",
    # T3·6 업무·CTA
    25: "T3·6", 29: "T3·6", 33: "T3·6", 37: "T3·6", 42: "T3·6",
    # T3·8 행정·CS
    16: "T3·8", 32: "T3·8", 36: "T3·8", 22: "T3·8", 44: "T3·8",
}


def _norm_priority(raw: str) -> str:
    if not raw:
        return Priority.UNSET
    raw = str(raw).strip()
    if raw.startswith("P0"): return Priority.P0
    if raw.startswith("P1"): return Priority.P1
    if raw.startswith("P2"): return Priority.P2
    if raw.startswith("P3"): return Priority.P3
    return Priority.UNSET


def _to_int(v):
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


class Command(BaseCommand):
    help = "포켓만프로젝트 엑셀(교육운영실 시트) → ops-catalog DB 1회 import"

    def add_arguments(self, parser):
        parser.add_argument("--xlsx", required=True, help="엑셀 파일 경로")
        parser.add_argument("--dry-run", action="store_true", help="DB 변경 없이 결과만 출력")

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise CommandError("openpyxl 미설치. requirements.txt 확인.") from e

        xlsx_path = Path(options["xlsx"])
        if not xlsx_path.exists():
            raise CommandError(f"파일 없음: {xlsx_path}")

        wb = load_workbook(xlsx_path, data_only=True)
        if "교육운영실" not in wb.sheetnames:
            raise CommandError("'교육운영실' 시트가 없습니다.")
        ws = wb["교육운영실"]

        dry = options["dry_run"]

        # 1) 카테고리 보장
        cat_objs: dict[str, Category] = {}
        for c in CATEGORIES:
            obj, created = Category.objects.update_or_create(
                code=c["code"],
                defaults={"tier": c["tier"], "num": c["num"], "title": c["title"], "note": c["note"]},
            )
            cat_objs[c["code"]] = obj
            if dry and created:
                self.stdout.write(f"  [dry] Category 생성: {c['code']}")
        if not dry:
            self.stdout.write(self.style.SUCCESS(f"카테고리 {len(cat_objs)}개 보장 완료"))

        # 2) 프로젝트 import
        created_count = 0
        updated_count = 0
        skipped = []

        with transaction.atomic():
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_idx == 1:
                    continue
                if not any(v is not None and str(v).strip() for v in row):
                    continue
                if row_idx not in ROW_TO_CAT:
                    skipped.append(row_idx)
                    continue

                priority, org, name, email, summary, blueprint, scope, manual_min, monthly_uses, _ = row

                defaults = {
                    "title": (str(summary) if summary else "").strip(),
                    "description": (str(blueprint) if blueprint else "").strip(),
                    "proposer": (str(name) if name else "").strip(),
                    "proposer_email": (str(email) if email else "").strip(),
                    "org": (str(org) if org else "").strip(),
                    "impact_scope": (str(scope) if scope else "").strip(),
                    "manual_minutes": _to_int(manual_min),
                    "monthly_uses": _to_int(monthly_uses),
                    "priority": _norm_priority(str(priority) if priority else ""),
                    "category": cat_objs[ROW_TO_CAT[row_idx]],
                    # 평가·진행 default
                    "difficulty": Difficulty.UNSET,
                    "deploy_intent": False,
                    "stage": Stage.S1,  # 미평가는 default 1단계
                }

                if dry:
                    exists = Project.objects.filter(source_id=row_idx).exists()
                    self.stdout.write(
                        f"  [dry] R{row_idx} ({'update' if exists else 'create'}): "
                        f"{defaults['proposer']} — {defaults['title'][:40]}"
                    )
                    continue

                obj, created = Project.objects.update_or_create(
                    source_id=row_idx,
                    defaults=defaults,
                )
                if created:
                    created_count += 1
                else:
                    updated_count += 1

            if dry:
                transaction.set_rollback(True)
                self.stdout.write(self.style.WARNING("dry-run — 모든 변경 롤백."))
                return

        if skipped:
            self.stdout.write(self.style.WARNING(f"매핑 없는 행 skip: {skipped}"))
        self.stdout.write(self.style.SUCCESS(
            f"✓ Project import 완료 — created={created_count}, updated={updated_count}"
        ))
